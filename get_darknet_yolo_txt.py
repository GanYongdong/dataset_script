##从保存的所有信息的xlsx文件中，读取需要的并转换成对应dakenet yolo4的txt label文件,所有图片目标都存在一个txt文件中

from openpyxl import load_workbook
import os
from pathlib import Path
import platform

if platform.system().lower() == 'windows':
    print("windows")
    xlsx_path = 'D:\\Research\\My_tamper_detect_dataset_generate\\dataset_tmp\\xlsx_all_info'
    txt_out_path = 'D:\\Research\\My_tamper_detect_dataset_generate\\dataset_tmp\\yolo_darknet_label_txt\\darknet_yolo_Label.txt'
elif platform.system().lower() == 'linux':
    print("linux")
    xlsx_path = "/media/ganyd/0C26A92126A90D30/Research/My_tamper_detect_dataset_generate/6dataset_irregular_png_06resize_3x3kernel_8class/xlsx_all_info"
    txt_out_path = '/media/ganyd/0C26A92126A90D30/Research/My_tamper_detect_dataset_generate/6dataset_irregular_png_06resize_3x3kernel_8class/yolo_darknet_label_txt/darknet_yolo_Label.txt'

# xlsx所在目录
path = xlsx_path

# 获取目录下所有目录与文件名称
fileNameList = os.listdir(path)

# 打开要写的txt文件，已有就清空，没有就创建
f = open(txt_out_path,
         'w')
f.close()
f = open(txt_out_path,
         'a')

for i in range(len(fileNameList)):

    xlsxFilePath = path + '/' + fileNameList[i]
    filePathTmp = Path(xlsxFilePath)
    if filePathTmp.is_dir():
        continue  # 如果这个是目录，不是文件则跳过

    # 转换成jpg目录
    picFilePath = fileNameList[i].split('.')[0]
    picFilePath = picFilePath + '.png'
    f.write(picFilePath)

    # 打开xlsx文件
    workbook = load_workbook(xlsxFilePath)  # 加载xlsx文件
    booksheet = workbook.active  # 获取活动的页
    # worksheet = workbook.get_sheet_by_name('Sheet1') #或者指定name获取sheet页
    rows = booksheet  # 获取sheet页的行数据
    columns = booksheet.columns  # 获取sheet页的列数据

    # 循环获取每个目标信息
    # [left,top,w,h,cx,cy,right,bottom,area, cx_rate, cy_rate, w_rate, h_rate]
    j = 0
    strWillWrite = ' '
    for row in rows:
        j = j + 1
        xmin = booksheet.cell(row=j, column=4).value
        ymin = booksheet.cell(row=j, column=5).value
        xmax = booksheet.cell(row=j, column=10).value
        ymax = booksheet.cell(row=j, column=11).value
        classId = booksheet.cell(row=j, column=1).value - 1
        strWillWrite = strWillWrite + str(xmin) + ' ' + str(ymin) + ' ' + str(xmax) + ' ' + str(ymax) + ' ' + str(
            classId) + ' '

    # 写txt
    f.write(strWillWrite)
    f.write('\n')
    # print(strWillWrite)
    if i % 100 == 0:
        print(i, '/', len(fileNameList))

# 关闭txt文件
f.close()
print('all finish!')
